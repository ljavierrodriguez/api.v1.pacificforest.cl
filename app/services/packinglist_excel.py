from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO
import os

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill


class PackingListExcelGenerator:
    def _build_sheet_title(self, guia_index: int, guia_despacho: str | None) -> str:
        base = f"Guia {guia_index}"
        if guia_despacho:
            base = f"{base} - {guia_despacho}"
        # Excel admite maximo 31 caracteres por nombre de hoja.
        return base[:31]

    def _get_logo_path(self):
        possible = [
            os.path.join(os.getcwd(), "app", "static", "logo_pacific_forest.png"),
            os.path.join(os.getcwd(), "app", "static", "logo.png"),
            os.path.join(os.getcwd(), "static", "logo_pacific_forest.png"),
            os.path.join(os.getcwd(), "static", "logo.png"),
        ]
        for path in possible:
            if os.path.exists(path):
                return path
        return None

    def _to_decimal(self, value) -> Decimal:
        if value is None:
            return Decimal("0")
        if isinstance(value, Decimal):
            return value
        try:
            return Decimal(str(value))
        except Exception:
            return Decimal("0")

    def _to_excel_value(self, value):
        if isinstance(value, Decimal):
            return float(value)
        if isinstance(value, (datetime, date)):
            return value
        return value

    def _get_detail_volume_m3(self, detail) -> Decimal:
        # VolumenLinea = (ancho * largo * espesor * piezas) / 1,000,000,000
        espesor = self._to_decimal(getattr(detail, "espesor", None))
        ancho = self._to_decimal(getattr(detail, "ancho", None))
        largo = self._to_decimal(getattr(detail, "largo", None))
        piezas = self._to_decimal(getattr(detail, "piezas", None))

        if espesor == 0 or ancho == 0 or largo == 0 or piezas == 0:
            return Decimal("0")

        return (ancho * largo * espesor * piezas) / Decimal("1000000000")

    def _write_logo(self, worksheet) -> None:
        logo_path = self._get_logo_path()
        if not logo_path:
            return

        try:
            logo_img = XLImage(logo_path)
            max_width = 220
            max_height = 80
            current_width = float(logo_img.width or 1)
            current_height = float(logo_img.height or 1)
            scale = min(max_width / current_width, max_height / current_height, 1.0)
            logo_img.width = int(current_width * scale)
            logo_img.height = int(current_height * scale)
            worksheet.add_image(logo_img, "A1")
            worksheet.row_dimensions[1].height = 64
            worksheet.row_dimensions[2].height = 16
        except Exception:
            return

    def _write_table_header(self, worksheet, row: int, header_fill: PatternFill) -> None:
        columns = [
            "OC",
            "Etiqueta",
            "Numero Paquetes",
            "Espesor",
            "Ancho",
            "Largo",
            "Piezas",
            "Vol m3",
            "Origen Detalle",
        ]

        for col_idx, column_name in enumerate(columns, start=1):
            cell = worksheet.cell(row=row, column=col_idx, value=column_name)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = header_fill

    def generate_excel(self, packing_list) -> BytesIO:
        workbook = Workbook()

        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        total_global_paquetes = 0
        total_global_volumen = Decimal("0")

        for guia_index, guia in enumerate(packing_list.guias, start=1):
            if guia_index == 1:
                worksheet = workbook.active
            else:
                worksheet = workbook.create_sheet()

            worksheet.title = self._build_sheet_title(guia_index, guia.guia_despacho)

            self._write_logo(worksheet)
            worksheet["C1"] = "Packing List"
            worksheet["C1"].font = Font(bold=True, size=16)

            header_rows = [
                ("Orden Compra ID", packing_list.orden_compra_id),
                ("Origen", packing_list.origen),
                ("Producto", packing_list.producto),
                ("Destino", packing_list.destino),
            ]

            row_cursor = 5
            for label, value in header_rows:
                worksheet.cell(row=row_cursor, column=1, value=label).font = Font(bold=True)
                worksheet.cell(row=row_cursor, column=2, value=self._to_excel_value(value))
                row_cursor += 1

            row_cursor += 1

            worksheet.cell(row=row_cursor, column=1, value=f"Guia #{guia_index}").font = Font(
                bold=True,
                size=12,
            )
            row_cursor += 1

            worksheet.cell(row=row_cursor, column=1, value="Guia despacho").font = Font(bold=True)
            worksheet.cell(row=row_cursor, column=2, value=guia.guia_despacho)
            row_cursor += 1

            worksheet.cell(row=row_cursor, column=1, value="Fecha despacho").font = Font(bold=True)
            worksheet.cell(row=row_cursor, column=2, value=self._to_excel_value(guia.fecha_despacho))
            row_cursor += 2

            self._write_table_header(worksheet, row_cursor, header_fill)
            row_cursor += 1

            total_guia_paquetes = 0
            total_guia_volumen = Decimal("0")

            for detail in guia.detalles:
                volumen_m3 = self._get_detail_volume_m3(detail)
                volumen_m3_display = volumen_m3.quantize(Decimal("0.001"), rounding=ROUND_HALF_UP)
                paquetes = detail.numero_pqts or 0

                total_guia_paquetes += paquetes
                total_guia_volumen += volumen_m3

                worksheet.cell(row=row_cursor, column=1, value=detail.oc)
                worksheet.cell(row=row_cursor, column=2, value=detail.etiqueta)
                worksheet.cell(row=row_cursor, column=3, value=paquetes)
                worksheet.cell(row=row_cursor, column=4, value=self._to_excel_value(detail.espesor))
                worksheet.cell(row=row_cursor, column=5, value=self._to_excel_value(detail.ancho))
                worksheet.cell(row=row_cursor, column=6, value=self._to_excel_value(detail.largo))
                worksheet.cell(row=row_cursor, column=7, value=detail.piezas)
                vol_cell = worksheet.cell(row=row_cursor, column=8, value=float(volumen_m3_display))
                vol_cell.number_format = "0.000"
                worksheet.cell(row=row_cursor, column=9, value=detail.origen_detalle)
                row_cursor += 1

            total_global_paquetes += total_guia_paquetes
            total_global_volumen += total_guia_volumen

            worksheet.cell(row=row_cursor, column=2, value="TOTAL GUIA").font = Font(bold=True)
            worksheet.cell(row=row_cursor, column=3, value=total_guia_paquetes).font = Font(bold=True)
            total_guia_vol_cell = worksheet.cell(
                row=row_cursor,
                column=8,
                value=float(total_guia_volumen.quantize(Decimal("0.001"), rounding=ROUND_HALF_UP)),
            )
            total_guia_vol_cell.font = Font(bold=True)
            total_guia_vol_cell.number_format = "0.000"

            for col_idx in (2, 3, 8):
                total_cell = worksheet.cell(row=row_cursor, column=col_idx)
                total_cell.fill = header_fill
                total_cell.alignment = Alignment(horizontal="center")

            column_widths = {
                "A": 20,
                "B": 24,
                "C": 20,
                "D": 14,
                "E": 14,
                "F": 14,
                "G": 14,
                "H": 14,
                "I": 22,
            }
            for column_letter, width in column_widths.items():
                worksheet.column_dimensions[column_letter].width = width

        summary_sheet = workbook.create_sheet("Resumen")
        self._write_logo(summary_sheet)
        summary_sheet["C1"] = "Packing List - Resumen"
        summary_sheet["C1"].font = Font(bold=True, size=16)

        summary_rows = [
            ("Orden Compra ID", packing_list.orden_compra_id),
            ("Origen", packing_list.origen),
            ("Producto", packing_list.producto),
            ("Destino", packing_list.destino),
            ("Total Guias", len(packing_list.guias)),
            ("Total Paquetes", total_global_paquetes),
            (
                "Total Volumen m3",
                float(total_global_volumen.quantize(Decimal("0.001"), rounding=ROUND_HALF_UP)),
            ),
        ]

        row_cursor = 5
        for label, value in summary_rows:
            summary_sheet.cell(row=row_cursor, column=1, value=label).font = Font(bold=True)
            value_cell = summary_sheet.cell(row=row_cursor, column=2, value=self._to_excel_value(value))
            if label == "Total Volumen m3":
                value_cell.number_format = "0.000"
            row_cursor += 1

        summary_sheet.column_dimensions["A"].width = 22
        summary_sheet.column_dimensions["B"].width = 24

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output
